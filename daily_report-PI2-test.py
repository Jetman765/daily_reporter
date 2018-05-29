import requests, json, openpyxl
import datetime as dt
import user
import pandas as pd
import numpy as np

def create_workbook():
	print('Creating workbook')
	t1 = dt.datetime.now()
	try:
		openpyxl.load_workbook("{} SENT2 Status-PI2-test.xlsx".format(create_date()))
		wb = openpyxl.load_workbook('{} SENT2 Status-PI2-test.xlsx'.format(create_date()))
		ws = wb.active
	except FileNotFoundError:
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Epic"

		ws["A1"].value = "Key"
		ws["B1"].value = "Summary"
		ws["C1"].value = "Title"
		ws["D1"].value = "Version"
		ws["E1"].value = "Story Points"
		ws["F1"].value = "TPO"
		ws["G1"].value = "Total Child Story Points"

	print('Creating workbook took {}'.format(dt.datetime.now() - t1))
	return wb

def create_date():
	date = "{year}-{month}-{day}".format(year=dt.datetime.now().year,
									month=dt.datetime.now().month if dt.datetime.now().month >= 10 else "0"+str(dt.datetime.now().month),
									day = dt.datetime.now().day if dt.datetime.now().day >= 10 else "0"+str(dt.datetime.now().day))
	return date

def save_workbook(wb):
	wb.save("{} SENT2 Status-PI2-test.xlsx".format(create_date()))

def edit_workbook(wb, dic):  #, epic):
	print('Starting edit_workbook')
	start = dt.datetime.now()
	ws = wb['Epic']

	row = 2
	for i in range(len(dic['issues'])):

		# Skip all deferred
		if dic['issues'][i]['fields']['status']['name'] == 'Deferred':
			continue

		if dic['issues'][i]['key'] == 'SENT2-12':
			continue

		ws["A{}".format(row)].value = dic['issues'][i]['key']
		ws["B{}".format(row)].value = dic['issues'][i]['fields']['summary']
		ws["C{}".format(row)].value = dic['issues'][i]['key'] + ' ' + dic['issues'][i]['fields']['summary']
		ws["E{}".format(row)].value = dic['issues'][i]['fields'].get('customfield_11213','')

		row += 1


	ws["C{}".format(row)].value = "Security Roll up"
	ws["G{}".format(row)].value = '=SUMIF($B$2:$B${row},"Security*",G$2:G${row})'.format(row=row)
	ws["H{}".format(row)].value = '=SUMIF($B$2:$B${row},"Security*",H$2:H${row})'.format(row=row)
	ws["I{}".format(row)].value = '=SUMIF($B$2:$B${row},"Security*",I$2:I${row})'.format(row=row)
	ws["J{}".format(row)].value = '=SUMIF($B$2:$B${row},"Security*",J$2:J${row})'.format(row=row)
	ws["K{}".format(row)].value = '=SUMIF($B$2:$B${row},"Security*",K$2:K${row})'.format(row=row)
	ws["D{}".format(row)].value = "SENT 1.0"

	mid1 = dt.datetime.now()
	print('About to collect_stories. Took {}'.format(mid1 - start))

	collect_stories(wb)

	mid2 = dt.datetime.now()
	print('collecting stories took {}'.format(mid2-mid1))

	if len(ws._tables) == 0:
		ws['H1'].value = 'Closed'
		ws['I1'].value = 'Work Done'
		ws['J1'].value = 'In Progress'
		ws['K1'].value = 'No Progress'

		tab = openpyxl.worksheet.table.Table(displayName="Epic",ref="A1:K{}".format(row))
		ws.add_table(tab)

def get_filter(filter):
	start = dt.datetime.now()
	r = json.loads(requests.get("https://{jira}/rest/api/2/filter/{filter}".format(filter=filter,jira=user.company_jira),auth=(user.username,user.password)).content)
	print('getting filter json took {}'.format(dt.datetime.now() - start))
	return r['jql']

def get_epics(jql_raw):
	start = dt.datetime.now()
	print('starting get_epics')
	open_paren, close_paren = jql_raw.find('('), jql_raw.find(')')
	jql_raw = jql_raw[open_paren+1:close_paren]
	epics = jql_raw.replace('%2C',',')
	new_jql = "https://{jira}/rest/api/2/search?jql=key+in+({epics})".format(epics=epics,jira=user.company_jira)
	new_jql += "&fields=summary,customfield_11213,customfield=11701,status,assignee&maxResults=200"

	jira_dic = json.loads(requests.get(new_jql, auth=(user.username,user.password)).content)

	print('ending  get_epics. took {}'.format(dt.datetime.now() - start))
	return jira_dic

def get_sprints(raw_sprint):
	l = []
	for i in range(len(raw_sprint)):
		st = raw_sprint[i].find('name=')
		en = raw_sprint[i][st:].find(',')
		l.append(raw_sprint[i][st+5:st+en])
	return l

def collect_stories(wb):
	print(wb.sheetnames)

	if "Story" not in wb.sheetnames:
		ws2 = wb.create_sheet(title="Story")
	else:
		ws2 = wb['Story']

	ws2["A1"].value = "Key"
	ws2["B1"].value = "Summary"
	ws2["C1"].value = "Story Points"	#customfield_11213
	ws2["D1"].value = "Sprint"			#customfield_11701
	ws2["E1"].value = "Team/s"			#customfield_14400
	ws2["F1"].value = "Epic Link"		#customfield_13300
	ws2["G1"].value = "Status"			
	ws2["H1"].value = "Resolution"
	ws2["I1"].value = "Original Story Points"
	ws2["J1"].value = "Type"

	ws2["K1"].value = "No Progress"
	ws2["L1"].value = "In Progress"
	ws2["M1"].value = "Work Done"
	ws2["N1"].value = "Closed"

	pre_url = dt.datetime.now()
	
	print("about to get json for stories")
	url = "https://{jira}/rest/api/2/search?jql={filter}&maxResults=3000".format(filter=get_filter(108698),jira=user.company_jira)
	url += "&fields=summary,customfield_11213,customfield_11701,customfield_14400,customfield_13300,status,resolution,issuetype&expand=issues,changelog"
	dic = json.loads(requests.get(url,auth=(user.username,user.password)).content)

	post_url = dt.datetime.now()
	print('getting the json took {}'.format(post_url - pre_url))

	teams_dic = {
		48065: "Common Services Team 2 - Darcy",
		48066: "Common Services Team 1 - Ranji",
		48067: "Common Services Team 2 - Jing",
		48068: "Common Services Team 1 - Sabari",
		48069: "IOT Team 1 - Saji",
		48070: "Software Team 1 - Martin",
		48071: "Data Team 1 - Sowmya",
		48072: "Data Team 2 - Arvind",
		48073: "Common Testing Team 1 - Duverney"
	}

	row = 2
	pre_print = dt.datetime.now()
	for i in range(len(dic['issues'])):
		# # Skip all rejected
		if dic['issues'][i]['fields']['resolution']:
				if dic['issues'][i]['fields']['resolution']['name'] == 'Rejected':
					continue

		# # Skip all deferred
		if dic['issues'][i]['fields']['status']['name'] == 'Deferred':
			continue

		ws2["A{}".format(row)].value = dic['issues'][i]['key']
		ws2["B{}".format(row)].value = dic['issues'][i]['fields']['summary']
		ws2["C{}".format(row)].value = dic['issues'][i]['fields'].get('customfield_11213','')

		if dic['issues'][i]['fields'].get('customfield_11701',''):
			ws2["D{}".format(row)].value = ', '.join(get_sprints(dic['issues'][i]['fields']['customfield_11701']))
		
		team = dic['issues'][i]['fields'].get('customfield_14400','')
		if team != None:
			ws2["E{}".format(row)].value = teams_dic.get(int(team[0]), "No Team")

		ws2["F{}".format(row)].value = dic['issues'][i]['fields'].get('customfield_13300','')
		ws2["G{}".format(row)].value = dic['issues'][i]['fields']['status']['name']
		if dic['issues'][i]['fields']['resolution'] != None:
			ws2["H{}".format(row)].value = dic['issues'][i]['fields']['resolution']['name']

		sp = []
		
		# Uncomment this if you want the original story points.
		# for j in range(dic['issues'][i]['changelog']['total']):
		# 	for k in range(len(dic['issues'][i]['changelog']['histories'][j]['items'])):
		# 		if dic['issues'][i]['changelog']['histories'][j]['items'][k]['field'] == 'Story Points':
		# 			sp.append(dic['issues'][i]['changelog']['histories'][j]['items'][k]['fromString'])
		# if len(sp) > 0 and sp[0] != None:
		# 	ws2["I{}".format(row)].value = float(sp[0])
		# else:
		# 	ws2["I{}".format(row)].value = dic['issues'][i]['fields'].get('customfield_11213','')

		ws2["I{}".format(row)].value = '...'

		ws2["J{}".format(row)].value = dic['issues'][i]['fields']['issuetype']['name']

		ws2["K{}".format(row)].value = '=IF(OR($L{}>0,$M{}>0,$N{}>0),0,$C{})'.format(row, row, row, row)
		ws2["L{}".format(row)].value = '=IF($G{}="In Progress",$C{},0)'.format(row,row)
		ws2["M{}".format(row)].value = '=IF(AND($H{}="Fixed",$G{}<>"Closed"),$C{},0)'.format(row,row,row)
		ws2["N{}".format(row)].value = '=IF($G{}="Closed",$C{},0)'.format(row,row)

		row += 1

	print('printing stories to worksheet took {}'.format(dt.datetime.now() - pre_print))

	if len(ws2._tables) == 0:
		ws2['O1'].value = "Title"
		ws2['P1'].value = "Version"
		tab2 = openpyxl.worksheet.table.Table(displayName="Story",ref="A1:P{}".format(row-1))
		ws2.add_table(tab2)


def make_epic_pivot_worksheet(wb):
	from openpyxl.utils.dataframe import dataframe_to_rows

	if "Epic Pivot" not in wb.sheetnames:
		ws = wb.create_sheet(title="Epic Pivot")
	else:
		ws = wb['Epic Pivot']

	epic_pt = make_pivot()
	
	for r in dataframe_to_rows(epic_pt, header=True):
		ws.append(r)

	ws.delete_rows(2, 1)

def make_pivot(sprint=None):
	if sprint == None:
		excel = pd.read_excel("{} SENT2 Status-PI2.xlsx".format(create_date()), sheet_name="Story")
		column_order = ['Closed', 'Work Done', 'In Progress', 'No Progress']
		pt = pd.pivot_table(excel, index='Title', aggfunc=np.sum)
		pt = pt.reindex(column_order, axis=1)
	else:
		excel = pd.read_excel("{} SENT2 Status-PI2.xlsx".format(create_date()), sheet_name="Story")
		sprint = excel[excel['Sprint']==sprint]
		column_order = ['Closed', 'Work Done', 'In Progress', 'No Progress']
		pt = pd.pivot_table(sprint, index='Title', aggfunc=np.sum)
		pt = pt.reindex(column_order, axis=1)

	return pt

def make_sprint_pivot_worksheet(wb):
	from openpyxl.utils.dataframe import dataframe_to_rows

	if "Sprint Pivot" not in wb.sheetnames:
		ws = wb.create_sheet(title="Sprint Pivot")
	else:
		ws = wb['Sprint Pivot']

	sprint_pt = make_pivot('*18.Q2.2.1 (Ends 6/4)')
	
	for r in dataframe_to_rows(sprint_pt, header=True):
		ws.append(r)	

	ws.delete_rows(2, 1)

def make_chart(ws):
	from openpyxl.chart import BarChart, Reference

	chart = BarChart()
	chart.type= 'bar'
	chart.style = 10
	chart.title = "Epic PI Progress"
	chart.grouping = 'stacked'
	
	data = Reference(ws, min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column)
	chart.add_data(data)

	ws.add_chart(chart,"F1")

def post_edit(wb):
	print('Editing workbook after dumping data')
	ws = wb["Epic"]

	ws['G{}'.format(2)].value = '\'=SUM(Epic[@[Closed]:[No Progress]])'
	max_epic_row = 0
	for row in range(2, ws.max_row):
		ws['H{}'.format(row)].value = '=SUMIF(Story[[#Headers],[#Data],[Epic Link]],[Key],Story[[#Headers],[Closed]])'
		ws['I{}'.format(row)].value = '=SUMIF(Story[[#Headers],[#Data],[Epic Link]],[Key],Story[[#Headers],[Work Done]])'
		ws['J{}'.format(row)].value = '=SUMIF(Story[[#Headers],[#Data],[Epic Link]],[Key],Story[[#Headers],[In Progress]])'
		ws['K{}'.format(row)].value = '=SUMIF(Story[[#Headers],[#Data],[Epic Link]],[Key],Story[[#Headers],[No Progress]])'
		max_epic_row = row

	ws = wb["Story"]

	for row in range(2, ws.max_row+1):
		epic_key = ws['F{}'.format(row)].value
		ws['O{}'.format(row)].value = '=IF(ISNUMBER(SEARCH("Security",INDEX(Epic!$C$2:$C${t_row},MATCH("{k_row}",Epic!$A$2:$A${t_row},0)))),"Security Roll up",INDEX(Epic!$C$2:$C${t_row},MATCH("{k_row}",Epic!$A$2:$A${t_row},0)))'.format(t_row = max_epic_row, k_row = epic_key)
		ws['P{}'.format(row)].value = '=IF([@Title]="Security Roll up","SENT 1.0",VLOOKUP([@[Epic Link]],Epic[#Data],4,))'

	make_epic_pivot_worksheet(wb)
	make_sprint_pivot_worksheet(wb)

	save_workbook(wb)


if __name__ == "__main__":
	wb = create_workbook()
	fil = get_filter(108698)
	epics = get_epics(fil)
	edit_workbook(wb, epics)
	save_workbook(wb)
	post_edit(wb)